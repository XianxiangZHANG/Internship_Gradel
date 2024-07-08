from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django.urls import path
from django import forms
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from django.db.models import Subquery, OuterRef, Count

from .calculator import PartCalculator, BushingCalculator, InterfaceCalculator, LinkCalculator
from web import models
from database.settings import BASE_DIR
import django_filters

class PartFilter(django_filters.FilterSet):
    project = django_filters.ModelChoiceFilter(queryset=models.Project.objects.all())
    partName = django_filters.CharFilter(field_name='partName', lookup_expr='icontains')
    resin = django_filters.ModelChoiceFilter(queryset=models.Resin.objects.all())
    fiber = django_filters.ModelChoiceFilter(queryset=models.Fiber.objects.all())
    valid = django_filters.BooleanFilter(field_name='valid')  

    class Meta:
        model = models.Part
        fields = ['project', 'partName', 'resin', 'fiber', 'valid']
   

class PartFilterValid(django_filters.FilterSet):
    project = django_filters.ModelChoiceFilter(queryset=models.Project.objects.all())
    partName = django_filters.CharFilter(field_name='partName', lookup_expr='icontains')
    resin = django_filters.ModelChoiceFilter(queryset=models.Resin.objects.all())
    fiber = django_filters.ModelChoiceFilter(queryset=models.Fiber.objects.all())
    

    class Meta:
        model = models.Part
        fields = ['project', 'partName', 'resin', 'fiber']
   

def get_filtered_parts(filter_class, request, valid=None):
    parts = None
    message = "No part to display. Please use the filter to load data."

    if valid is not None:
        queryset = models.Part.objects.filter(valid=valid)
    else:
        queryset = models.Part.objects.all()

    part_filter = filter_class(request.GET, queryset=queryset)
    if any(request.GET.values()):
        parts = part_filter.qs.select_related('fiber', 'resin', 'project').prefetch_related('link_set', 'bushing_set', 'interface_set')
        if not parts.exists():
            message = "No data found."
    elif 'filter' in request.GET:
        parts = part_filter.qs.select_related('fiber', 'resin', 'project').prefetch_related('link_set', 'bushing_set', 'interface_set')
        if not parts.exists():
            message = "No data found."
    
    if parts:
        for part in parts:
            calculator = PartCalculator(part)
            part.calculated_properties = {
                'fiberDensity': calculator.fiber_density,
                'resinDensity': calculator.resin_density,
                'windingDensity': calculator.winding_density,
                'totalMassLink': calculator.total_mass_link(),
                'totalMassAccumulation': calculator.total_mass_accumulation(),
                'totalMassWinding': calculator.total_mass_winding(),
                'totalMassBushing': calculator.total_mass_bushing(),
                'totalMassStructure': calculator.total_mass_structure(),
                'totalFiberLength': calculator.total_fiber_length(),
                'totalFiberMass': calculator.total_fiber_mass(),
                'totalResinMass': calculator.total_resin_mass(),
            }

    return part_filter, parts, message

def part_list(request):
    part_filter, parts, message = get_filtered_parts(PartFilter, request)
    return render(request, 'part/part_list.html', {'filter': part_filter, 'parts': parts, 'message': message})

def part_list_doc(request):
    part_filter, parts, message = get_filtered_parts(PartFilter, request)
    return render(request, 'part/part_list_doc.html', {'filter': part_filter, 'parts': parts, 'message': message})

def part_valid(request):
    part_filter, parts, message = get_filtered_parts(PartFilterValid, request, valid=True)
    return render(request, 'part/part_valid.html', {'filter': part_filter, 'parts': parts, 'message': message})

def part_valid_doc(request):
    part_filter, parts, message = get_filtered_parts(PartFilterValid, request, valid=True)
    return render(request, 'part/part_valid_doc.html', {'filter': part_filter, 'parts': parts, 'message':message})



def part_input(request):
    """ list of part """

    queryset = models.Part.objects.all().order_by("id")
    
    return render(request, 'part/part_input.html', {"queryset": queryset})

class PartModelForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName', 'resin', 'fiber',  'totalTowNumber', 'fiberVolumeRatio', 'fiberSectionCalc', 'fiberSectionAcc',
                  'defaultInterfaceHeight', 'defaultInterfaceIntDiam', 'defaultLinkType', 'defaultLinkDefined', 
                   'additionalMass', 
                  'projectImage','valid'
                #   'part_gh', 'part_mod', 'part_csv', 'part_rs', 'part_log', 'part_mp4', 'part_jpg'
                  ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}

class PartModelAddForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName', 'resin','fiber',  'totalTowNumber', 'fiberVolumeRatio', 'fiberSectionCalc', 'fiberSectionAcc',
                  'defaultInterfaceHeight', 'defaultInterfaceIntDiam', 'defaultLinkType', 'defaultLinkDefined', 
                   'additionalMass', 
                  'projectImage',
                 ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}

class PartModelImageForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName',
                  'projectImage',
                 ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


class PartModelDocForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName',
                  'part_gh', 'part_mod', 'part_csv', 'part_rs', 'part_log', 'part_mp4', 'part_jpg',
                 ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


def part_add(request):
    if request.method == "GET":
        form = PartModelAddForm()
        return render(request, 'part/part_add.html', {"form": form})

    form = PartModelAddForm(request.POST,request.FILES)
    if not form.is_valid():

        print("Form is not valid:", form.errors)
        return render(request, 'part/part_add.html', {"form": form})


    # form.save()
    part = form.save(commit=False)
    part.user = models.User.objects.filter(id=request.info_dict['id']).first() 
    part.save()
    return redirect('/part/list/')


class PartEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName', 'resin', 'fiber', 'totalTowNumber', 'fiberVolumeRatio', 'fiberSectionCalc', 'fiberSectionAcc',
                  'defaultInterfaceHeight', 'defaultInterfaceIntDiam', 'defaultLinkType', 'defaultLinkDefined', 
                  'additionalMass', 
                  'projectImage','valid'
                #   'part_gh', 'part_mod', 'part_csv', 'part_rs', 'part_log', 'part_mp4', 'part_jpg'
                  ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}

class PartEditModelDocForm(forms.ModelForm):
    class Meta:
        model = models.Part
        fields = ['project', 'partName',
                  'part_gh', 'part_mod', 'part_csv', 'part_rs', 'part_log', 'part_mp4', 'part_jpg',
                 ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}

def part_edit(request, aid):
    part_object = models.Part.objects.filter(id=aid).first()

    if request.method == "GET":
        form = PartEditModelForm(instance=part_object)
        return render(request, 'part/part_form.html', {"form": form})

    form = PartEditModelForm(request.POST,request.FILES, instance=part_object)
    if not form.is_valid():
        return render(request, 'part/part_form.html', {"form": form})

    # form.save()
    part = form.save(commit=False)
    part.user = models.User.objects.filter(id=request.info_dict['id']).first() 
    part.save()

    return redirect('/part/list/')



def part_edit_doc(request, aid):
    part_object = models.Part.objects.filter(id=aid).first()

    if request.method == "GET":
        form = PartEditModelDocForm(instance=part_object)
        return render(request, 'part/part_form.html', {"form": form})

    form = PartEditModelDocForm(request.POST,request.FILES, instance=part_object)
    if not form.is_valid():
        return render(request, 'part/part_form.html', {"form": form})

    # form.save()
    part = form.save(commit=False)
    part.user = models.User.objects.filter(id=request.info_dict['id']).first() 
    part.save()

    return redirect('/part/list-doc/')

def part_delete(request):
    aid = request.GET.get("aid")
    # models.Part.objects.filter(id=aid).delete()
    part = models.Part.objects.filter(id=aid).first()
    if part:
        part.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        part.delete()

    return JsonResponse({"status": True})

def handle_uploaded_file_part(f):
    # load Excel document
    wb = load_workbook(filename=f, data_only=True)
    if 'CoverPage' not in wb.sheetnames:
        return {'error': 'CoverPage sheet not found'}
    if 'Input' not in wb.sheetnames:
        return {'error': 'Input sheet not found'}
    if 'Output' not in wb.sheetnames:
        return {'error': 'Output sheet not found'}
    
    sheetCP = wb['CoverPage']
    sheetIn = wb['Input']
    sheetOut = wb['Output']

    project_data = {}
    for row in sheetCP.iter_rows():
        for cell in row:
            if cell.value == "Program : ":
                project_data['program'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Customer : ":
                project_data['customer'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Project No:":
                project_data['projectNo'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']

    part_data = {}
    
    part_data['project_id'] = models.Project.objects.filter(projectName=project_data['projectName']).first().id
    # print(part_data['project_id'])
    part_data['partName'] = project_data['projectName']


    for row in sheetIn.iter_rows():
        for cell in row:
            if cell.value == "Interface Height [mm]":
                part_data['defaultInterfaceHeight'] = int(sheetIn.cell(row=cell.row, column=cell.column + 1).value)
            elif cell.value == "Interface Int. Diam [mm]":
                part_data['defaultInterfaceIntDiam'] = int(sheetIn.cell(row=cell.row, column=cell.column + 1).value)
            elif cell.value == "Resin name":
                part_data['resinName'] = sheetIn.cell(row=cell.row, column=cell.column+1 ).value
            elif cell.value == "Fiber name":
                part_data['fiberName'] = sheetIn.cell(row=cell.row, column=cell.column+1 ).value
            elif cell.value == "Fiber volume ratio [%]":
                part_data['fiberVolumeRatio'] = float(sheetIn.cell(row=cell.row, column=cell.column+1 ).value)
            elif cell.value == "Fiber Section calc. [mm²]":
                part_data['fiberSectionCalc'] = sheetIn.cell(row=cell.row, column=cell.column+1 ).value
            elif cell.value == "Fiber Section acc. [mm²]":
                part_data['fiberSectionAcc'] = sheetIn.cell(row=cell.row, column=cell.column+1 ).value
            elif cell.value == "Link (Element) Type" and sheetIn.cell(row=cell.row+1, column=cell.column ).value == "Link defined by":
                part_data['defaultLinkType'] = sheetIn.cell(row=cell.row, column=cell.column+1 ).value
            elif cell.value == "Link defined by":
                part_data['defaultLinkDefined'] = sheetIn.cell(row=cell.row, column=cell.column + 1).value
            elif cell.value == "Number of spools":
                part_data['totalTowNumber'] = sheetIn.cell(row=cell.row, column=cell.column + 1).value

    resin_name = str(part_data['resinName'])
    matching_resin = models.Resin.objects.filter(resin__icontains=resin_name).first()
    if matching_resin is not None:
        part_data['resin_id'] = matching_resin.id
    else:
        part_data['resin_id'] = None
    fiber_name = str(part_data['fiberName'])
    matching_fiber = models.Fiber.objects.filter(grade__icontains=fiber_name).first()
    if matching_fiber is not None:
        part_data['fiber_id'] = matching_fiber.id
    else:
        part_data['fiber_id'] = None

    for row in sheetOut.iter_rows():
        for cell in row:
            if cell.value == "Number of links":
                part_data['numberLink'] = int(sheetOut.cell(row=cell.row, column=cell.column + 1).value)
            elif cell.value == "Number of bushings":
                part_data['numberBushing'] = int(sheetOut.cell(row=cell.row, column=cell.column + 1).value)
            elif cell.value == "Total mass of links [g]":
                part_data['totalMassLink'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total mass of accumulation [g]":
                part_data['totalMassAccumulation'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total mass of winding [g]":
                part_data['totalMassWinding'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total mass of bushings [g]":
                part_data['totalMassBushing'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Additional masses [g]":
                part_data['additionalMass'] =  sheetOut.cell(row=cell.row, column=cell.column + 1).value 
            elif cell.value == "Total mass of structure [g]":
                part_data['totalMassStructure'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total fiber length [m]":
                part_data['totalFiberLength'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total fiber mass [kg]":
                part_data['totalFiberMass'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            elif cell.value == "Total resin mass [g]":
                part_data['totalResinMass'] = round(sheetOut.cell(row=cell.row, column=cell.column + 1).value, 2)
            
            

    return part_data

def upload_file_part(request):
    if request.method == 'POST' and 'file' in request.FILES:
        part_data = handle_uploaded_file_part(request.FILES['file'])
        if 'error' in part_data:
            return JsonResponse(part_data, status=400)
        return JsonResponse(part_data)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def format_value(value):
    return str(value) if value else "--"

def download_parts_pdf(request):
    f = PartFilterValid(request.GET, queryset=models.Part.objects.filter(valid=True))
    parts = f.qs

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="parts.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 40, "Parts List")
    p.setFont("Helvetica-Bold", 10)
    p.drawString(150, height - 40, "\"--\" means the value is None")

    # p.setFont("Helvetica", 10)
    x = 50
    xx = 200
    y = height - 90
    l = 20
    i = 1
    for index, part in enumerate(parts):
        p.line(x/2, y +15 , width - x/2, y+15)
        p.setFont("Helvetica-Bold", 10)
        p.drawString(x, y, "Project Name")
        p.drawString(x, y-l, "Part Name")
        p.drawString(x, y-2*l, "Default Interface Height")
        p.drawString(x, y-3*l, "Default Interface Int Diam")
        p.drawString(x, y-4*l, "Default Link Type")
        p.drawString(x, y-5*l, "Default Link Defined")
        p.drawString(x, y-6*l, "Number of Links")
        p.drawString(x, y-7*l, "Number of Bushings")
        p.drawString(x, y-8*l, "Total Mass Link")
        p.drawString(x, y-9*l, "Total Mass Accumulation")
        p.drawString(x, y-10*l, "Total Mass Winding")
        p.drawString(x, y-11*l, "Total Mass Bushing")
        p.drawString(x, y-12*l, "Additional Mass")
        p.drawString(x, y-13*l, "Total Mass Structure")
        p.drawString(x, y-14*l, "Total Fiber Length")
        p.drawString(x, y-15*l, "Total Fiber Mass")
        p.drawString(x, y-16*l, "Total Resin Mass")
        p.drawString(x, y-17*l, "Project Image")
        p.setFont("Helvetica", 10)
        p.drawString(xx, y, format_value(part.project.projectName))
        p.drawString(xx, y-l, format_value(part.partName))
        p.drawString(xx, y-2*l, format_value(part.defaultInterfaceHeight))
        p.drawString(xx, y-3*l, format_value(part.defaultInterfaceIntDiam))
        p.drawString(xx, y-4*l, format_value(part.defaultLinkType))
        p.drawString(xx, y-5*l, format_value(part.defaultLinkDefined))
        p.drawString(xx, y-6*l, format_value(part.numberLink))
        p.drawString(xx, y-7*l, format_value(part.numberBushing))
        p.drawString(xx, y-8*l, format_value(part.totalMassLink))
        p.drawString(xx, y-9*l, format_value(part.totalMassAccumulation))
        p.drawString(xx, y-10*l, format_value(part.totalMassWinding))
        p.drawString(xx, y-11*l, format_value(part.totalMassBushing))
        p.drawString(xx, y-12*l, format_value(part.additionalMass))
        p.drawString(xx, y-13*l, format_value(part.totalMassStructure))
        p.drawString(xx, y-14*l, format_value(part.totalFiberLength))
        p.drawString(xx, y-15*l, format_value(part.totalFiberMass))
        p.drawString(xx, y-16*l, format_value(part.totalResinMass))

        if part.projectImage:

            image_path = str(BASE_DIR) + part.projectImage.url
            print(image_path)
            p.drawImage(image_path, xx, y-17*l - 300, width = 200, height = 300, preserveAspectRatio=True, mask='auto' )
        else:
            p.drawString(xx, y-17*l, "--")
        y -= 380
        if index < len(parts) - 1:  
            p.showPage()
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, height - 40, "Parts List")
            p.setFont("Helvetica-Bold", 10)
            p.drawString(150, height - 40, "\"--\" means the value is None")
            y = height - 90
          

    p.showPage()
    p.save()
    return response