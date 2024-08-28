from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django import forms
from django.forms import modelformset_factory
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph

from .calculator import PartCalculator, BushingCalculator, InterfaceCalculator, LinkCalculator
from web import models
import django_filters

class InterfaceFilter(django_filters.FilterSet):
    class Meta:
        model = models.Interface
        fields = {
            'project': ['exact'], 
            'part': ['exact'],
            'interfaceName': ['icontains'],
        }

def get_filtered_interfaces(filter_class, request, valid=None):
    interfaces = None
    message = "No interface to display. Please use the filter to load data."

    if valid is not None:
        queryset = models.Interface.objects.filter(part__valid=valid)
    else:
        queryset = models.Interface.objects.all()

    interface_filter = filter_class(request.GET, queryset=queryset)
    if any(request.GET.values()):
        interfaces = interface_filter.qs
        if not interfaces.exists():
            message = "No data found."
    elif 'filter' in request.GET:
        interfaces = interface_filter.qs
        if not interfaces.exists():
            message = "No data found."

    if interfaces:
        for interface in interfaces:
            calculator = InterfaceCalculator(interface)
            interface.calculated_properties = {
                'totalLink': calculator.total_link(),
                'totalArm': calculator.total_arm(),
                'totalSection': calculator.total_section(),
                'extDiameter': calculator.ext_diameter(),
                'accMass': calculator.acc_mass(),
                'safetyFactor': calculator.safety_factor(),
            }

    return interface_filter, interfaces, message

def interface_list(request):
    interface_filter, interfaces, message = get_filtered_interfaces(InterfaceFilter, request)
    return render(request, 'interface/interface_list.html', {'filter': interface_filter, 'interfaces': interfaces, 'message': message})

def interface_valid(request):
    interface_filter, interfaces, message = get_filtered_interfaces(InterfaceFilter, request, valid=True)
    return render(request, 'interface/interface_valid.html', {'filter': interface_filter, 'interfaces': interfaces, 'message': message})


def interface_input(request):
    """ list of interface """

    # [obj,]
    queryset = models.Interface.objects.all().order_by("id")
  
    return render(request, 'interface/interface_input.html', {"queryset": queryset})

   
class InterfaceModelForm(forms.ModelForm):
    class Meta:
        model = models.Interface
        fields = ['interfaceName', 'height', 'intDiameter',
                'finODiam', 'finAccSection', 
                'interfaceCenterX', 'interfaceCenterY', 'interfaceCenterZ',
                'directionVectorX', 'directionVectorY', 'directionVectorZ',
                'divisionStep',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field_object in self.fields.items():
            field_object.widget.attrs = {"class": "form-control"}


class ProjectPartForm(forms.Form):
    project = forms.ModelChoiceField(queryset=models.Project.objects.all(), required=True)
    part = forms.ModelChoiceField(queryset=models.Part.objects.all(), required=True)
    

InterfaceFormSet = modelformset_factory(
    models.Interface,
    form=InterfaceModelForm,
    extra=1  
)

def interface_add_multiple(request):
    projects = models.Project.objects.all()
    parts = models.Part.objects.all()
    formset = InterfaceFormSet(queryset=models.Interface.objects.none())
    project_part_form = ProjectPartForm()

    if request.method == 'POST':
        formset = InterfaceFormSet(request.POST)
        project_part_form = ProjectPartForm(request.POST)

        if formset.is_valid() and project_part_form.is_valid(): 
            instances = formset.save(commit=False)
            project = project_part_form.cleaned_data['project']
            part = project_part_form.cleaned_data['part']
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.project = project
                instance.part = part
                instance.save()
            return redirect('/interface/list/')  # Replace with your redirect URL
        else:
            # Handle formset or project_part_form validation errors here
            print(formset.errors)
            print(project_part_form.errors)

    return render(request, 'interface/interface_add_multiple.html', {
        'projects': projects,
        'parts': parts,
        'formset': formset,
        'project_part_form': project_part_form,
        'interfaceError': "Interface name is Required",
    })



def interface_add(request):
    if request.method == "GET":
        form = InterfaceModelForm()
        return render(request, 'interface/interface_form.html', {"form": form})

    form = InterfaceModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'interface/interface_form.html', {"form": form})


    # form.save()
    interface = form.save(commit=False)
    interface.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    interface.save()
    return redirect('/interface/list/')

def interface_modify_multiple(request):
    # Get filtered data
    interface_filter = InterfaceFilter(request.GET, queryset=models.Interface.objects.all())
    
    # Define form set
    InterfaceFormSet = modelformset_factory(models.Interface, form=InterfaceModelForm, extra=0)
    
    formset = None
    message = "No interface to display. Please use the filter to load data."

    if request.method == 'POST':
        formset = InterfaceFormSet(request.POST)
        if formset.is_valid():
            # formset.save()
            instances = formset.save(commit=False)
            
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.save()
            return redirect('/interface/list/')
    else:
        if any(request.GET.values()):
            formset = InterfaceFormSet(queryset=interface_filter.qs)
            if not interface_filter.qs.exists():
                message = "No data found."
        elif 'filter' in request.GET:
            formset = InterfaceFormSet(queryset=interface_filter.qs)
            if not interface_filter.qs.exists():
                message = "No data found."

    return render(request, 'interface/interface_modify_multiple.html', {
        'filter': interface_filter,
        'formset': formset,
        'interfaceError': "Interface name is Required",
        'message': message,
    })

class InterfaceEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Interface
        fields = ['project', 'part', 'interfaceName', 'height', 'intDiameter', 
                'finODiam', 'finAccSection',
                'interfaceCenterX', 'interfaceCenterY', 'interfaceCenterZ',
                'directionVectorX', 'directionVectorY', 'directionVectorZ',
                'divisionStep',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.fields['part'].queryset = models.Part.objects.all()
        self.fields['project'].queryset = models.Project.objects.all()
        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


def interface_edit(request, aid):
    interface_object = models.Interface.objects.filter(id=aid).first()

    if request.method == "GET":
        form = InterfaceEditModelForm(instance=interface_object)
        return render(request, 'interface/interface_form.html', {"form": form})

    form = InterfaceEditModelForm(instance=interface_object, data=request.POST)
    if not form.is_valid():
        return render(request, 'interface/interface_form.html', {"form": form})

    # form.save()
    interface = form.save(commit=False)
    interface.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    interface.save()

    return redirect('/interface/list/')


def interface_delete(request):
    aid = request.GET.get("aid")
    # models.Interface.objects.filter(id=aid).delete()
    interface = models.Interface.objects.filter(id=aid).first()
    if interface:
        interface.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        interface.delete()

    return JsonResponse({"status": True})

def handle_uploaded_file_interface(f):
    # load Excel document
    wb = load_workbook(filename=f, data_only=True)
    if 'CoverPage' not in wb.sheetnames:
        return {'error': 'CoverPage sheet not found'}
    if 'Input' not in wb.sheetnames:
        return {'error': 'Input sheet not found'}
    if 'Output' not in wb.sheetnames:
        return {'error': 'Output sheet not found'}
    if 'Bushing Table' not in wb.sheetnames:
        return {'error': 'Bushing Table sheet not found'}
    
    sheetCP = wb['CoverPage']
    sheetIn = wb['Input']
    sheetOut = wb['Output']
    sheetBT = wb['Bushing Table']

    # get data
    interface_data = {}
    interfaceName = {}
    height = {}
    intDiameter = {}
    totalLink = {}
    totalArm = {}
    totalSection = {}
    extDiameter = {}
    accMass = {}
    finODiam = {}
    finAccSection = {}

    safetyFactor = {}

    interfaceCenterX = {}
    interfaceCenterY = {}
    interfaceCenterZ = {}
    # directionVectorX = {}
    # directionVectorY = {}
    # directionVectorZ = {}

    # divisionStep = {}
    
    project_data = {}
    for row in sheetCP.iter_rows():
        for cell in row:
            if cell.value == "Program : ":
                project_data['program'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Customer : ":
                project_data['customer'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Project No:":
                project_data['projectNo'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']

    
    interface_data['project_id'] = models.Project.objects.filter(projectName=project_data['projectName']).first().id
    interface_data['part_id'] = models.Part.objects.filter(partName=project_data['projectName']).first().id


    # for row in sheetCP.iter_rows():
    #     for cell in row:
    #         if cell.value == "Program : ":
    #             # interface_data['project_projectName'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    #             interface_data['project_id'] = models.Project.objects.filter(projectName=sheetCP.cell(row=cell.row, column=cell.column + 2).value).first().id
               

    for row in sheetOut.iter_rows():
        for cell in row:
            if cell.value == "Number of bushings":
                numberBushing = int(sheetOut.cell(row=cell.row, column=cell.column + 1).value)
                # print(type(numberBushing))
    
    numberInterface = 0
    for row in sheetIn.iter_rows():
        for cell in row:
            if cell.value == "# of int.":
                for i in range(0, numberBushing):
                    numberInterface += int(sheetIn.cell(row=cell.row+1+i, column=cell.column ).value)
                    # print(numberInterface)

    for row in sheetBT.iter_rows():
        for cell in row:
            if cell.value == "Interface" and cell.row > 3:
                for i in range(0, numberInterface):
                    # print(cell.row, cell.column)
                    interfaceName[i] = sheetBT.cell(row=cell.row, column=cell.column + 1 + i ).value
                    print(interfaceName[i])
            elif cell.value == "Height [mm]":
                for i in range(0, numberInterface):
                    height[i] = int(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value)
                    # print(height[i])
            elif cell.value == "Int. Diameter [mm]":
                for i in range(0, numberInterface):
                    intDiameter[i] = int(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value)
                    # print(intDiameter[i])
            elif cell.value == "Total Link":
                for i in range(0, numberInterface):
                    totalLink[i] = int(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value)
                    # print(totalLink[i])
            elif cell.value == "Total Arm":
                for i in range(0, numberInterface):
                    totalArm[i] = int(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value)
                    # print(totalArm[i])
            elif cell.value == "Total Section [mm²]":
                for i in range(0, numberInterface):
                    totalSection[i] = round(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value, 2)
                    # print(totalSection[i])
            elif cell.value == "Ext. Diameter [mm]":
                for i in range(0, numberInterface):
                    extDiameter[i] = round(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value, 2)
                    # print(extDiameter[i])
            elif cell.value == "Acc. Mass [g]":
                for i in range(0, numberInterface):
                    accMass[i] = round(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value, 2)
                    # print(accMass[i])
            elif cell.value == "Fin. O. Diam. [mm]":
                for i in range(0, numberInterface):
                    finODiam[i] = sheetBT.cell(row=cell.row, column=cell.column+1+i ).value
                    # print(finODiam[i])
            elif cell.value == "Fin. Acc. section [mm²]":
                for i in range(0, numberInterface):
                    finAccSection[i] = sheetBT.cell(row=cell.row, column=cell.column+1+i ).value
                    # print(finAccSection[i])
            elif cell.value == "Safety Factor [%]":
                for i in range(0, numberInterface):
                    safetyFactor[i] = int(sheetBT.cell(row=cell.row, column=cell.column+1+i ).value * 100)
                    # print(safetyFactor[i])
            elif cell.value == "X":
                for i in range(0, numberInterface):
                    interfaceCenterX[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(interfaceCenterX[i])
            elif cell.value == "Y":
                for i in range(0, numberInterface):
                    interfaceCenterY[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(interfaceCenterY[i])
            elif cell.value == "Z":
                for i in range(0, numberInterface):
                    interfaceCenterZ[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(interfaceCenterZ[i])
            # elif cell.value == "":
            #     for i in range(0, numberInterface):
            #         directionVectorX[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
            #         # print(directionVectorX[i])
            # elif cell.value == "":
            #     for i in range(0, numberInterface):
            #         directionVectorY[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
            #         # print(directionVectorY[i])
            # elif cell.value == "":
            #     for i in range(0, numberInterface):
            #         directionVectorZ[i] = sheetBT.cell(row=cell.row+1+i, column=cell.column ).value
            #         # print(directionVectorZ[i])
    
    interface =  {
        'numberInterface':numberInterface,
        'interface_data':interface_data,
        'interfaceName':interfaceName,
        'height':height,
        'intDiameter':intDiameter,
        'totalLink':totalLink,
        'totalArm':totalArm,
        'totalSection':totalSection,
        'extDiameter':extDiameter,
        'accMass':accMass,
        'finODiam':finODiam,
        'finAccSection':finAccSection,
        'safetyFactor':safetyFactor,
        'interfaceCenterX':interfaceCenterX,
        'interfaceCenterY':interfaceCenterY,
        'interfaceCenterZ':interfaceCenterZ,
    }
    return interface

def upload_file_interface(request):
    if request.method == 'POST' and 'file' in request.FILES:
        interface = handle_uploaded_file_interface(request.FILES['file'])
        print(interface)
        if 'error' in interface:
            return JsonResponse(interface, status=400)
        return JsonResponse(interface)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def draw_rotated_header(canvas, x, y, text, width, height, angle):
    canvas.saveState()
    canvas.translate(x, y)
    canvas.rotate(angle)
    p = Paragraph(text, getSampleStyleSheet()['Normal'])
    p.wrapOn(canvas, height, width)  # width and height are swapped because of rotation
    p.drawOn(canvas, 0, 0)
    canvas.restoreState()


def format_value(value):
        return str(value) if value else "--"

def download_interfaces_pdf(request):
    interface_filter, interfaces, message = get_filtered_interfaces(InterfaceFilter, request, valid=True)
    # f = InterfaceFilter(request.GET, queryset=models.Interface.objects.filter(part__valid=True))
    # interfaces = f.qs

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="interfaces.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    styles = getSampleStyleSheet()

    if not interfaces:
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(30, height - 40, "Interfaces List")
        p.setFont("Helvetica-Bold", 10)
        p.drawString(150, height - 40, "\"--\" means the value is None")
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 60, "No Interfaces found. Please adjust your filters and try again.")

    else:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(30, height - 40, "Interfaces List")
        p.setFont("Helvetica-Bold", 10)
        p.drawString(150, height - 40, "\"--\" means the value is None")
        p.setDash() 
        p.setLineWidth(1)
        p.line(30, height - 50, width - 30, height - 50)
        p.setFont("Helvetica-Bold", 10)
        y_start = height - 70
        x_offset = 30
        y_offset = 20  
        
        headers = ["Interface Name", "Height [mm]", "Int. Diameter [mm]", "Total Link",
                "Total Arm", "Total Section [mm²]", "Ext. Diameter [mm]", "Acc. Mass [g]",
                "Fin. O. Diam. [mm]", "Fin. Acc. section [mm²]", "Safety Factor [%]", "Interface's center X [mm]",
                "Interface's center Y [mm]", "Interface's center Z [mm]", "Direction vector X [mm]", "Direction vector Y [mm]",
                "Direction vector Z [mm]","Division Step"]
        column_widths = [30, 18, 18, 18,
                        18, 35, 35, 35,
                        35, 35, 35, 38,
                        38, 38, 38, 38, 
                        38, 18]

        current_project = None
        current_part = None

        p.setFont("Helvetica", 7)
        y = y_start

        for index, interface in enumerate(interfaces):
            next_is_title = (index + 1 < len(interfaces)) and (
                interfaces[index + 1].project.projectName != interface.project.projectName or 
                interfaces[index + 1].part.partName != interface.part.partName
            )

            if interface.project.projectName != current_project or interface.part.partName != current_part:
                if current_project is not None:
                    y -= 10  
                    p.setDash() 
                    p.setLineWidth(1)
                    p.line(x_offset, y, width - x_offset, y)

                current_project = interface.project.projectName
                current_part = interface.part.partName
                y -= y_offset

                p.setFont("Helvetica-Bold", 10)
                p.drawString(x_offset, y, "Project Name: " + current_project)
                y -= y_offset
                p.drawString(x_offset, y, "Part      Name: " + current_part)
                y -= y_offset
                p.setDash() 
                p.setLineWidth(1)
                p.line(x_offset, y, width - x_offset, y)
                y -= y_offset

                for i, header in enumerate(headers):
                    draw_rotated_header(p, x_offset + sum(column_widths[:i]) + column_widths[i] / 2, y - 5 * y_offset, header, 200,200,  90)
                y -= 6 * y_offset


                p.setFont("Helvetica", 9)

            
            p.setDash(1, 2) 
            p.setLineWidth(1)
            values = [
                format_value(interface.interfaceName), format_value(interface.height), format_value(interface.intDiameter), format_value(interface.calculated_properties.get('totalLink')),
                format_value(interface.calculated_properties.get('totalArm')), format_value(f"{interface.calculated_properties.get('totalSection', 0):.2f}"), format_value(f"{interface.calculated_properties.get('extDiameter', 0):.2f}"), format_value(f"{interface.calculated_properties.get('accMass', 0):.2f}"),
                format_value(interface.finODiam), format_value(interface.finAccSection), format_value(interface.calculated_properties.get('safetyFactor')), format_value(interface.interfaceCenterX),
                format_value(interface.interfaceCenterY), format_value(interface.interfaceCenterZ), format_value(interface.directionVectorX), format_value(interface.directionVectorY),
                format_value(interface.directionVectorZ), format_value(interface.divisionStep)
            ]

            p.drawString(x_offset, y, values[0])
            for i in range(1, len(values)):
                p.line(x_offset + sum(column_widths[:i]) - 4, y + y_offset / 2, x_offset + sum(column_widths[:i]) - 4, y - y_offset / 2)
                p.drawString(x_offset + sum(column_widths[:i]), y, values[i])

            y -= y_offset
            if next_is_title:
                p.showPage()
                p.setFont("Helvetica-Bold", 12)
                p.drawString(30, height - 40, "Interfaces List")
                p.setFont("Helvetica-Bold", 10)
                p.drawString(150, height - 40, "\"--\" means the value is None")
                p.setDash() 
                p.setLineWidth(1)
                p.line(30, height - 50, width - 30, height - 50)
                p.setFont("Helvetica-Bold", 10)
                y = height - 70
                current_project = None
                current_part = None 
            else:
                if y < 100 :
                    p.showPage()
                    p.setFont("Helvetica-Bold", 12)
                    p.drawString(30, height - 40, "Interfaces List (continued)")
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(200, height - 40, "\"--\" means the value is None")
                    p.setDash() 
                    p.setLineWidth(1)
                    p.line(30, height - 50, width - 30, height - 50)
                    p.setFont("Helvetica-Bold", 10)
                    y = height - 70
                    current_project = None
                    current_part = None 

    p.showPage()
    p.save()
    return response
