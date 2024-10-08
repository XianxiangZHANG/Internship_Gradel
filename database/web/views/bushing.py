from typing import Type
from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django import forms
from django.forms import modelformset_factory
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Frame
from reportlab.lib import colors

from .calculator import PartCalculator, BushingCalculator, InterfaceCalculator, LinkCalculator
from web import models
import django_filters

class BushingFilter(django_filters.FilterSet):
    project = django_filters.ModelChoiceFilter(queryset=models.Project.objects.all())
    part = django_filters.ModelChoiceFilter(queryset=models.Part.objects.all())
    bushingName = django_filters.CharFilter(field_name='bushingName', lookup_expr='icontains')
    
    class Meta:
        model = models.Bushing
        fields = ['project', 'part', 'bushingName']


def get_filtered_bushings(filter_class, request, valid=None):
    bushings = None
    message = "No bushing to display. Please use the filter to load data."

    if valid is not None:
        queryset = models.Bushing.objects.filter(part__valid=valid)
    else:
        queryset = models.Bushing.objects.all()

    bushing_filter = filter_class(request.GET, queryset=queryset)
    if any(request.GET.values()):
        bushings = bushing_filter.qs
        if not bushings.exists():
            message = "No data found."
    elif 'filter' in request.GET:
        bushings = bushing_filter.qs
        if not bushings.exists():
            message = "No data found."

    if bushings:
        for bushing in bushings:
            calculator = BushingCalculator(bushing)
            bushing.calculated_properties = {
                'accOnBushing': calculator.acc_on_bushing(),
                'totalBushingMass': calculator.total_bushing_mass(),
            }

    return bushing_filter, bushings, message

def bushing_list(request):
    bushing_filter, bushings, message = get_filtered_bushings(BushingFilter, request)
    return render(request, 'bushing/bushing_list.html', {'filter': bushing_filter, 'bushings': bushings, 'message': message})

def bushing_valid(request):
    bushing_filter, bushings, message = get_filtered_bushings(BushingFilter, request, valid=True)
    return render(request, 'bushing/bushing_valid.html', {'filter': bushing_filter, 'bushings': bushings, 'message': message})



def bushing_input(request):
    """ list of bushing """

    # [obj,]
    queryset = models.Bushing.objects.all().order_by("id")
    
    return render(request, 'bushing/bushing_input.html', {"queryset": queryset})

   
class BushingModelForm(forms.ModelForm):
    class Meta:
        model = models.Bushing
        fields = ['bushingName', 'numberInterface', 'bushingDrawNb', 'bushingMass', ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field_object in self.fields.items():
            field_object.widget.attrs = {"class": "form-control"}

class BushingForm(forms.ModelForm):
    class Meta:
        model = models.Bushing
        fields = ['project', 'part', 'bushingName', 'numberInterface', 'bushingDrawNb',  'bushingMass',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field_object in self.fields.items():
            field_object.widget.attrs = {"class": "form-control"}


class ProjectPartForm(forms.Form):
    project = forms.ModelChoiceField(queryset=models.Project.objects.all(), required=True)
    part = forms.ModelChoiceField(queryset=models.Part.objects.all(), required=True)

BushingFormSet = modelformset_factory(
    models.Bushing,
    form=BushingModelForm,
    extra=1  # The number of additional forms added during initialization
)

BushingMultFormSet = modelformset_factory(
    models.Bushing,
    form=BushingForm,
    extra=1  # The number of additional forms added during initialization
)

def bushing_add_multiple(request):
    projects = models.Project.objects.all()
    parts = models.Part.objects.all()
    formset = BushingFormSet(queryset=models.Bushing.objects.none())
    project_part_form = ProjectPartForm()

    if request.method == 'POST':
        formset = BushingFormSet(request.POST)
        project_part_form = ProjectPartForm(request.POST)
        # bushingName = request.POST.get('bushingName')
       
            
        if formset.is_valid() and project_part_form.is_valid():
            instances = formset.save(commit=False)
            project = project_part_form.cleaned_data['project']
            part = project_part_form.cleaned_data['part']
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.project = project
                instance.part = part
                instance.save()
            return redirect('/bushing/list/')  # Replace with your redirect URL
        else:
            # Handle formset or project_part_form validation errors here
            print(formset.errors)
            print(project_part_form.errors)
            
    return render(request, 'bushing/bushing_add_multiple.html', {
        'projects': projects,
        'formset': formset,
        'parts': parts,
        'project_part_form': project_part_form,
        'bushingError': "Bushing name is Required",
    })


def bushing_modify_multiple(request):
    # Get filtered data
    bushing_filter = BushingFilter(request.GET, queryset=models.Bushing.objects.all())
    
    # Define form set
    BushingFormSet = modelformset_factory(models.Bushing, form=BushingModelForm, extra=0)
    
    formset = None
    message = "No bushing to display. Please use the filter to load data."

    if request.method == 'POST':
        # print("post")
        formset = BushingFormSet(request.POST)
        # print(formset)
        if formset.is_valid():
            instances = formset.save(commit=False)
            
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.save()
            return redirect('/bushing/list/')
    else:
        if any(request.GET.values()):
            formset = BushingFormSet(queryset=bushing_filter.qs)
            if not bushing_filter.qs.exists():
                message = "No data found."
        elif 'filter' in request.GET:
            formset = BushingFormSet(queryset=bushing_filter.qs)
            if not bushing_filter.qs.exists():
                message = "No data found."

    return render(request, 'bushing/bushing_modify_multiple.html', {
        'filter': bushing_filter,
        'formset': formset,
        'bushingError': "Bushing name is Required",
        'message': message,
    })



def bushing_add(request):
    if request.method == "GET":
        form = BushingModelForm()
        return render(request, 'bushing/bushing_form.html', {"form": form})

    form = BushingModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'bushing/bushing_form.html', {"form": form})


    # save -> DB
    # form.save()
    bushing = form.save(commit=False)
    bushing.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    bushing.save()
    return redirect('/bushing/list/')


class BushingEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Bushing
        fields = ['project', 'part', 'bushingName', 'numberInterface', 'bushingDrawNb', 'bushingMass', ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.fields['part'].queryset = models.Part.objects.all()
        self.fields['project'].queryset = models.Project.objects.all()
        for name, filed_object in self.fields.items():
            # print(name, filed_object)
            filed_object.widget.attrs = {"class": "form-control"}


def bushing_edit(request, aid):
    bushing_object = models.Bushing.objects.filter(id=aid).first()

    if request.method == "GET":
        form = BushingEditModelForm(instance=bushing_object)
        return render(request, 'bushing/bushing_form.html', {"form": form})

    form = BushingEditModelForm(instance=bushing_object, data=request.POST)
    if not form.is_valid():
        return render(request, 'bushing/bushing_form.html', {"form": form})

    # form.save()
    bushing = form.save(commit=False)
    bushing.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    bushing.save()

    return redirect('/bushing/list/')


def bushing_delete(request):
    aid = request.GET.get("aid")
    # models.Bushing.objects.filter(id=aid).delete()
    bushing = models.Bushing.objects.filter(id=aid).first()
    if bushing:
        bushing.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        bushing.delete()

    return JsonResponse({"status": True})

def bushing_delete_mult(request):
    aid = request.GET.get("aid")
    if not aid:
        return JsonResponse({"status": False, "error": "ID cannot be empty"})

    try:
        bushing = models.Bushing.objects.get(id=aid)
        # bushing.delete()
        if bushing:
            bushing.user = models.User.objects.filter(id=request.info_dict['id']).first()  
            bushing.delete()
        return JsonResponse({"status": True})
    except models.Bushing.DoesNotExist:
        return JsonResponse({"status": False, "error": "Bushing not found"})
    

def handle_uploaded_file_bushing(f):
    # load Excel document
    wb = load_workbook(filename=f, data_only=True)
    if 'CoverPage' not in wb.sheetnames:
        return {'error': 'CoverPage sheet not found'}
    if 'Input' not in wb.sheetnames:
        return {'error': 'Input sheet not found'}
    if 'Output' not in wb.sheetnames:
        return {'error': 'Output sheet not found'}
    
    sheetCP = wb['CoverPage']
    sheetIn = wb['Input']
    sheetOut = wb['Output']

    # get data
    bushing_data = {}
    
    bushing_data_numberInterface = {}
    bushing_data_position = {}
    bushing_data_draw = {}
    bushing_data_acc = {}
    bushing_data_mass = {}
    bushing_data_totalMass = {}

    project_data = {}
    for row in sheetCP.iter_rows():
        for cell in row:
            if cell.value == "Program : ":
                project_data['program'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Customer : ":
                project_data['customer'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Project No:":
                project_data['projectNo'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']

    
    bushing_data['project_id'] = models.Project.objects.filter(projectName=project_data['projectName']).first().id
    bushing_data['part_id'] = models.Part.objects.filter(partName=project_data['projectName']).first().id

    # for row in sheetCP.iter_rows():
    #     for cell in row:
    #         if cell.value == "Program : ":
    #             # bushing_data['project_projectName'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    #             bushing_data['project_id'] = models.Project.objects.filter(projectName=sheetCP.cell(row=cell.row, column=cell.column + 2).value).first().id
    #             # bushing_data['part_id'] = models.Part.objects.filter(partName=sheetCP.cell(row=cell.row+1, column=cell.column + 2).value).first().id
    #             # print(bushing_data['project_id'])

    for row in sheetOut.iter_rows():
        for cell in row:
            if cell.value == "Number of bushings":
                numberBushing = int(sheetOut.cell(row=cell.row, column=cell.column + 1).value)
                # print(type(numberBushing))
            
    for row in sheetIn.iter_rows():
        for cell in row:
            if cell.value == "# of int.":
                for i in range(0, numberBushing):
                    bushing_data_numberInterface[i] = int(sheetIn.cell(row=cell.row+1+i, column=cell.column ).value)
                    # print(bushing_data_numberInterface[i])

    for row in sheetOut.iter_rows():
        for cell in row:
            if cell.value == "Bush. Position":
                for i in range(0, numberBushing):
                    bushing_data_position[i] = sheetOut.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(bushing_data_position[i])
            elif cell.value == "Bush. Draw. nb.":
                for i in range(0, numberBushing):
                    bushing_data_draw[i] = sheetOut.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(bushing_data_draw[i])
            elif cell.value == "Acc. on bush. [g]":
                for i in range(0, numberBushing):
                    bushing_data_acc[i] = round(sheetOut.cell(row=cell.row+1+i, column=cell.column ).value, 2)
                    # print(bushing_data_acc[i])
            elif cell.value == "Bushing mass [g]":
                for i in range(0, numberBushing):
                    bushing_data_mass[i] = round(sheetOut.cell(row=cell.row+1+i, column=cell.column ).value, 2)
                    # print(bushing_data_mass[i])
            elif cell.value == "Total bush. mass [g]":
                for i in range(0, numberBushing):
                    bushing_data_totalMass[i] = round(sheetOut.cell(row=cell.row+1+i, column=cell.column ).value, 2)
                    # print(bushing_data_totalMass[i])
    
    bushing =  {
        'bushing_data': bushing_data,
        'numberBushing': numberBushing,
        'bushing_data_draw': bushing_data_draw,
        'bushing_data_numberInterface': bushing_data_numberInterface,
        'bushing_data_position': bushing_data_position,
        'bushing_data_acc': bushing_data_acc,
        'bushing_data_mass': bushing_data_mass,
        'bushing_data_totalMass': bushing_data_totalMass,
    }
    return bushing

def upload_file_bushing(request):
    if request.method == 'POST' and 'file' in request.FILES:
        bushing = handle_uploaded_file_bushing(request.FILES['file'])
        # print(bushing)
        if 'error' in bushing:
            return JsonResponse(bushing, status=400)
        return JsonResponse(bushing)
    return JsonResponse({'error': 'Invalid request'}, status=400)

def format_value(value):
        return str(value) if value else "--"


def download_bushings_pdf(request):
    bushing_filter, bushings, message = get_filtered_bushings(BushingFilter, request, valid=True)
    # f = BushingFilter(request.GET, queryset=models.Bushing.objects.filter(part__valid=True))
    # bushings = f.qs

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bushings.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    styles = getSampleStyleSheet()

    if not bushings:
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(30, height - 40, "Bushings List")
        p.setFont("Helvetica-Bold", 10)
        p.drawString(150, height - 40, "\"--\" means the value is None")
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 60, "No bushings found. Please adjust your filters and try again.")

    else:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(30, height - 40, "Bushings List")
        p.setFont("Helvetica-Bold", 10)
        p.drawString(150, height - 40, "\"--\" means the value is None")
        p.setDash() 
        p.setLineWidth(1)
        p.line(30, height - 50, width - 30, height - 50)
        p.setFont("Helvetica-Bold", 10)
        y_start = height - 70
        x_offset = 30
        y_offset = 20  
        headers = ["Bushing Name", "Number of Interface", "Bushing Draw. Nb", "Acc. On Bushing[g]", "Bushing Mass[g]", "Total Bushing Mass[g]"]
        column_widths = [80, 80, 100, 90, 90, 90]  

        current_project = None
        current_part = None

        p.setFont("Helvetica", 10)
        y = y_start

        for index, bushing in enumerate(bushings):
            next_is_title = (index + 1 < len(bushings)) and (
                bushings[index + 1].project.projectName != bushing.project.projectName or 
                bushings[index + 1].part.partName != bushing.part.partName
            )

            if bushing.project.projectName != current_project or bushing.part.partName != current_part:
                if current_project is not None:
                    y -= 10  
                    p.setDash() 
                    p.setLineWidth(1)
                    p.line(x_offset, y, width - x_offset, y)

                current_project = bushing.project.projectName
                current_part = bushing.part.partName
                y -= y_offset

                p.setFont("Helvetica-Bold", 10)
                p.drawString(x_offset, y, "Project Name: " + current_project)
                y -= y_offset
                p.drawString(x_offset, y, "Part      Name: " + current_part)
                y -= y_offset
                p.setDash(1, 2) 
                p.setLineWidth(1)
                p.line(x_offset, y, width - x_offset, y)
                y -= y_offset

                for i, header in enumerate(headers):
                    frame = Frame(x_offset + sum(column_widths[:i]), y-y_offset, column_widths[i], 2*y_offset, showBoundary=0)
                    paragraph = Paragraph(header, styles['Normal'])
                    frame.addFromList([paragraph], p)
                y -= 2*y_offset

                p.setFont("Helvetica", 10)

            p.drawString(20+x_offset, y, format_value(bushing.bushingName))
            p.drawString(20+x_offset + sum(column_widths[:1]), y, format_value(bushing.numberInterface))
            p.drawString(20+x_offset + sum(column_widths[:2]), y, format_value(bushing.bushingDrawNb))
            p.drawString(20+x_offset + sum(column_widths[:3]), y, format_value(bushing.calculated_properties.get('accOnBushing')))
            p.drawString(20+x_offset + sum(column_widths[:4]), y, format_value(bushing.bushingMass))
            p.drawString(20+x_offset + sum(column_widths[:5]), y, format_value(bushing.calculated_properties.get('totalBushingMass')))

            y -= y_offset
            if next_is_title:
                
                p.showPage()
                p.setFont("Helvetica-Bold", 12)
                p.drawString(30, height - 40, "Bushings List")
                p.setFont("Helvetica-Bold", 10)
                p.drawString(150, height - 40, "\"--\" means the value is None")
                p.setDash(1, 2) 
                p.setLineWidth(1)
                p.line(30, height - 50, width - 30, height - 50)
                p.setFont("Helvetica-Bold", 10)
                y = height - 70
                current_project = None
                current_part = None 
            else:
                if y < 50 :
                    p.showPage()
                    p.setFont("Helvetica-Bold", 12)
                    p.drawString(30, height - 40, "Bushings List (continued)")
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(250, height - 40, "\"--\" means the value is None")
                    p.setDash(1, 2) 
                    p.setLineWidth(1)
                    p.line(30, height - 50, width - 30, height - 50)
                    p.setFont("Helvetica-Bold", 10)
                    y = height - 70
                    current_project = None
                    current_part = None 

    p.showPage()
    p.save()
    return response