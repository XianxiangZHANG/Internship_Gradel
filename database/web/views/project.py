from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django import forms
from openpyxl import load_workbook
from django.forms.widgets import DateInput
from datetime import date, datetime
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from web import models
import django_filters

class ProjectFilter(django_filters.FilterSet):
    projectName = django_filters.CharFilter(field_name='projectName', lookup_expr='icontains')
    program = django_filters.CharFilter(field_name='program', lookup_expr='icontains')
    equipment = django_filters.CharFilter(field_name='equipment', lookup_expr='icontains')
    customer = django_filters.CharFilter(field_name='customer', lookup_expr='icontains')
    valid = django_filters.BooleanFilter(field_name='valid')  

    class Meta:
        model = models.Project
        fields = ['projectName', 'program', 'equipment', 'customer','valid']


def project_list(request):
    """ list of project """
    projects = None
    project_filter = ProjectFilter(request.GET, queryset=models.Project.objects.all())

    message = "No project to display. Please use the filter to load data."

    if any(request.GET.values()):
        projects = project_filter.qs
        message = "No data found."
    elif 'filter' in request.GET:
        projects = project_filter.qs
        message = "No data found."
    
    return render(request, 'project/project_list.html', {'filter': project_filter, 'projects': projects, 'message':message})


class ProjectFilterValid(django_filters.FilterSet):
    projectName = django_filters.CharFilter(field_name='projectName', lookup_expr='icontains')
    program = django_filters.CharFilter(field_name='program', lookup_expr='icontains')
    equipment = django_filters.CharFilter(field_name='equipment', lookup_expr='icontains')
    customer = django_filters.CharFilter(field_name='customer', lookup_expr='icontains')
    
    class Meta:
        model = models.Project
        fields = ['projectName', 'program', 'equipment', 'customer',]


def project_valid(request):
    """ list of project """

    projects = None
    project_filter = ProjectFilterValid(request.GET, queryset=models.Project.objects.filter(valid=True))

    message = "No project to display. Please use the filter to load data."

    if any(request.GET.values()):
        projects = project_filter.qs
        message = "No data found."
    elif 'filter' in request.GET:
        projects = project_filter.qs
        message = "No data found."
    
    return render(request, 'project/project_valid.html', {'filter': project_filter, 'projects': projects, 'message':message})


class ProjectModelForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ['projectName', 'program', 'equipment', 'customer', 'projectNo', 'relativeDesign', 'structureDrawingNb', 'documentNb', 'revision', 'lastUpdate']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


def project_add(request):
    if request.method == "GET":
        form = ProjectModelForm()
        return render(request, 'project/project_add.html', {"form": form})

    form = ProjectModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'project/project_add.html', {"form": form})


    # form.save()
    project = form.save(commit=False)
    project.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    project.save()
    return redirect('/project/list/')
    

class CustomDateInput(DateInput):
    input_type = 'text'
    
    def __init__(self, **kwargs):
        kwargs['format'] = '%m/%d/%Y'
        super().__init__(**kwargs)
    
    def format_value(self, value):
        if value is None:
            return ''
        if isinstance(value, (datetime.date, datetime.datetime)):
            return value.strftime(self.format)
        return value

class ProjectEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ['projectName', 'program', 'equipment', 'customer', 'projectNo', 'relativeDesign', 'structureDrawingNb', 'documentNb', 'revision', 'lastUpdate',]
        widgets = {
            'lastUpdate': CustomDateInput(),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}

    

def project_edit(request, aid):
    project_object = models.Project.objects.filter(id=aid).first()

    if request.method == "GET":
        form = ProjectEditModelForm(instance=project_object)
        return render(request, 'project/project_form.html', {"form": form})

    form = ProjectEditModelForm(instance=project_object, data=request.POST)
    if not form.is_valid():
        return render(request, 'project/project_form.html', {"form": form})

    # form.save()
    project = form.save(commit=False)
    project.user = models.User.objects.filter(id=request.info_dict['id']).first()
    project.save()
    return redirect('/project/list/')


def project_delete(request):
    aid = request.GET.get("aid")
    # models.Project.objects.filter(id=aid).delete()
    project = models.Project.objects.filter(id=aid).first()
    if project:
        project.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        project.delete()

    return JsonResponse({"status": True})

class UploadFileForm(forms.Form):
    file = forms.FileField()


def handle_uploaded_file_project(f):
    # load Excel document
    wb = load_workbook(filename=f, data_only=True)
    if 'CoverPage' not in wb.sheetnames:
        return {'error': 'CoverPage sheet not found'}
    
    sheet = wb['CoverPage']

    # get data
    project_data = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Program : ":
                project_data['program'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Equipment : ":
                project_data['equipment'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Customer : ":
                project_data['customer'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Project No:":
                project_data['projectNo'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Relative design:":
                project_data['relativeDesign'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Structure drawing nb:":
                project_data['structureDrawingNb'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Document Nr. : ":
                project_data['documentNb'] = sheet.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Revision: ":
                project_data['revision'] = sheet.cell(row=cell.row, column=cell.column + 1).value
            elif cell.value == "Last update :":
                last_update = sheet.cell(row=cell.row, column=cell.column + 1).value
                if isinstance(last_update, (datetime.date, datetime.datetime)):
                    project_data['lastUpdate'] = last_update.strftime('%m/%d/%Y')
                else:
                    project_data['lastUpdate'] = last_update
    
    project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']
    print(project_data['projectName'])
    return project_data

def upload_file_project(request):
    if request.method == 'POST' and 'file' in request.FILES:
        project_data = handle_uploaded_file_project(request.FILES['file'])
        if 'error' in project_data:
            return JsonResponse(project_data, status=400)
        
        return JsonResponse(project_data)
    return JsonResponse({'error': 'Invalid request'}, status=400)

def format_value(value):
        return str(value) if value else "--"

def download_projects_pdf(request):
    
    f = ProjectFilterValid(request.GET, queryset=models.Project.objects.filter(valid=True))
    projects = f.qs

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="projects.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    if not projects:
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, height - 40, "Projects List")
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 60, "No projects found. Please adjust your filters and try again.")

    else:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, height - 40, "Projects List")
        p.setFont("Helvetica-Bold", 10)
        p.drawString(150, height - 40, "\"--\" means the value is None")

        # p.setFont("Helvetica", 10)
        x = 50
        xx = 200
        y = height - 90
        l = 20
        i = 1
        for index, project in enumerate(projects):
            p.line(x/2, y +15 , width - x/2, y+15)
            p.setFont("Helvetica-Bold", 10)
            p.drawString(x, y, "Project Name")
            p.drawString(x, y-l, "Program")
            p.drawString(x, y-2*l, "Equipment")
            p.drawString(x, y-3*l, "Customer")
            p.drawString(x, y-4*l, "Project No")
            p.drawString(x, y-5*l, "Relative Design")
            p.drawString(x, y-6*l, "Structure Drawing Nb")
            p.drawString(x, y-7*l, "Document Nb")
            p.drawString(x, y-8*l, "Revision")
            p.drawString(x, y-9*l, "Last Update")
            p.setFont("Helvetica", 10)
            p.drawString(xx, y, format_value(project.projectName))
            p.drawString(xx, y-l, format_value(project.program))
            p.drawString(xx, y-2*l, format_value(project.equipment))
            p.drawString(xx, y-3*l, format_value(project.customer))
            p.drawString(xx, y-4*l, format_value(project.projectNo))
            p.drawString(xx, y-5*l, format_value(project.relativeDesign))
            p.drawString(xx, y-6*l, format_value(project.structureDrawingNb))
            p.drawString(xx, y-7*l, format_value(project.documentNb))
            p.drawString(xx, y-8*l, format_value(project.revision))
            p.drawString(xx, y-9*l, format_value(project.lastUpdate.strftime("%m-%d-%Y") ))
            y -= 220
            if y < 250 and index < len(projects) - 1:
                p.showPage()
                p.setFont("Helvetica-Bold", 12)
                p.drawString(50, height - 40, "Projects List")
                p.setFont("Helvetica-Bold", 10)
                p.drawString(150, height - 40, "\"--\" means the value is None")
                y = height - 90
            

    p.showPage()
    p.save()
    return response