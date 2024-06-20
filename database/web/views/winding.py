from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django import forms
from django.forms import modelformset_factory
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from web import models
import django_filters

class WindingFilter(django_filters.FilterSet):
    class Meta:
        model = models.Winding
        fields = {
            'project': ['exact'], 
            'part': ['exact'],
        }

def winding_list(request):
    """ list of winding """

    winding_filter = WindingFilter(request.GET, queryset=models.Winding.objects.all())
    return render(request, 'winding/winding_list.html', {'filter': winding_filter})

def winding_valid(request):
    """ list of winding """

    winding_filter = WindingFilter(request.GET, queryset=models.Winding.objects.filter(part__valid=True))
    return render(request, 'winding/winding_valid.html', {'filter': winding_filter})

def winding_input(request):
    """ list of winding """

    # [obj,]
    queryset = models.Winding.objects.all().order_by("id")
  
    return render(request, 'winding/winding_input.html', {"queryset": queryset})

   
class WindingModelForm(forms.ModelForm):
    class Meta:
        model = models.Winding
        fields = ['link', 'interface1','interface2','interface3',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field_object in self.fields.items():
            field_object.widget.attrs = {"class": "form-control"}


class ProjectPartForm(forms.Form):
    project = forms.ModelChoiceField(queryset=models.Project.objects.all(), required=True)
    part = forms.ModelChoiceField(queryset=models.Part.objects.all(), required=True)
    # link = forms.ModelChoiceField(queryset=models.Link.objects.all(), required=True)
    

WindingFormSet = modelformset_factory(
    models.Winding,
    form=WindingModelForm,
    extra=1 
)

def winding_add_multiple(request):
    projects = models.Project.objects.all()
    formset = WindingFormSet(queryset=models.Winding.objects.none())
    project_part_form = ProjectPartForm()


    if request.method == 'POST':
        formset = WindingFormSet(request.POST)
        project_part_form = ProjectPartForm(request.POST)

        if formset.is_valid() and project_part_form.is_valid(): 
            print("valid")
            instances = formset.save(commit=False)
            project = project_part_form.cleaned_data['project']
            part = project_part_form.cleaned_data['part']
            # link = project_part_form.cleaned_data['link']
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.project = project
                instance.part = part
                # instance.link = link
                instance.save()
            return redirect('/winding/list/')  # Replace with your redirect URL
        else:
            # Handle formset or project_part_form validation errors here
            print(formset.errors)
            print(project_part_form.errors)

    return render(request, 'winding/winding_add_multiple.html', {
        'projects': projects,
        'formset': formset,
        'project_part_form': project_part_form,
        'linkError': "Link name is Required",
    })

def winding_modify_multiple(request):
    # Get filtered data
    winding_filter = WindingFilter(request.GET, queryset=models.Winding.objects.all())
    
    # Define form set
    WindingFormSet = modelformset_factory(models.Winding, form=WindingModelForm, extra=0)
    
    if request.method == 'POST':
        formset = WindingFormSet(request.POST)
        if formset.is_valid():
            instances = formset.save(commit=False)
            
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.save()
            return redirect('/winding/list/')
    else:
        formset = WindingFormSet(queryset=winding_filter.qs)

    return render(request, 'winding/winding_modify_multiple.html', {
        'filter': winding_filter,
        'formset': formset,
        'linkError': "Link name is Required",
    })

def winding_add(request):
    if request.method == "GET":
        form = WindingModelForm()
        return render(request, 'winding/winding_form.html', {"form": form})

    form = WindingModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'winding/winding_form.html', {"form": form})


    # form.save()
    winding = form.save(commit=False)
    winding.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    winding.save()
    return redirect('/winding/list/')


class WindingEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Winding
        fields = ['project', 'part', 'link', 'interface1','interface2','interface3',  ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.fields['part'].queryset = models.Part.objects.all()
        self.fields['project'].queryset = models.Project.objects.all()
        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


def winding_edit(request, aid):
    winding_object = models.Winding.objects.filter(id=aid).first()

    if request.method == "GET":
        form = WindingEditModelForm(instance=winding_object)
        return render(request, 'winding/winding_form.html', {"form": form})

    form = WindingEditModelForm(instance=winding_object, data=request.POST)
    if not form.is_valid():
        return render(request, 'winding/winding_form.html', {"form": form})

    # form.save()
    winding = form.save(commit=False)
    winding.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    winding.save()

    return redirect('/winding/list/')


def winding_delete(request):
    aid = request.GET.get("aid")
    # models.Winding.objects.filter(id=aid).delete()
    winding = models.Winding.objects.filter(id=aid).first()
    if winding:
        winding.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        winding.delete()
    
    return JsonResponse({"status": True})


def winding_delete_mult(request):
    aid = request.GET.get("aid")
    if not aid:
        return JsonResponse({"status": False, "error": "ID cannot be empty"})

    try:
        winding = models.Winding.objects.get(id=aid)
        # winding.delete()
        if winding:
            winding.user = models.User.objects.filter(id=request.info_dict['id']).first()  
            winding.delete()

        return JsonResponse({"status": True})
    except models.Winding.DoesNotExist:
        return JsonResponse({"status": False, "error": "Winding not found"})
    

# def handle_uploaded_file_winding(f):
#     # 加载Excel文件
#     wb = load_workbook(filename=f, data_only=True)
#     if 'CoverPage' not in wb.sheetnames:
#         return {'error': 'CoverPage sheet not found'}
   
    
#     sheetCP = wb['CoverPage']

#     numberLink = 0

#     link_data = {}
   
    
#     project_data = {}
#     for row in sheetCP.iter_rows():
#         for cell in row:
#             if cell.value == "Program : ":
#                 project_data['program'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
#             elif cell.value == "Customer : ":
#                 project_data['customer'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
#             elif cell.value == "Project No:":
#                 project_data['projectNo'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
#     project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']

    
#     link_data['project_id'] = models.Project.objects.filter(projectName=project_data['projectName']).first().id


#     for row in sheetLT.iter_rows():
#         for cell in row:
#             if cell.value == "Link":
#                 i = 1
#                 while sheetLT.cell(row=cell.row+i, column=cell.column ).value != None:
#                     numberLink += 1
#                     print(numberLink)
#                     i +=1
#                 print(numberLink)

#     for row in sheetLT.iter_rows():
#         for cell in row:
#             if cell.value == "Link":
#                 for i in range(0, numberLink):
#                     # print(cell.row, cell.column)
#                     linkName[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
#                     print(linkName[i])
#             elif cell.value == "Bushing Bi":
#                 for i in range(0, numberLink):
#                     interface1[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column).value
#                     # print(interface1[i])
#             elif cell.value == "Bushing Bj":
#                 for i in range(0, numberLink):
#                     interface2[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
#                     # print(interface2[i])
#             elif cell.value == "Sequence":
#                 for i in range(0, numberLink):
#                     sequence[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column).value
#                     # print(sequence[i])
#             elif cell.value == "Length [mm]":
#                 for i in range(0, numberLink):
#                     lengthLink[i] = round(sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value,2)
#                     # print(lengthLink[i])
#             elif cell.value == "Arm diam. [mm]":
#                 for i in range(0, numberLink):
#                     armDiam[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
#                     # print(armDiam[i])
#             elif cell.value == "Arm Sec. [mm²]":
#                 for i in range(0, numberLink):
#                     armSection[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
#                     # print(armSection[i])
#             elif cell.value == "Cycle #":
#                 for i in range(0, numberLink):
#                     cycle[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
#                     # print(cycle[i])
#             elif cell.value == "Fin. Arm Sec. [mm²]":
#                 for i in range(0, numberLink):
#                     finArmSection[i] = round(sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value,2)
#                     # print(finArmSection[i])
#             elif cell.value == "Fin. Arm diam. [mm]":
#                 for i in range(0, numberLink):
#                     finArmDiam[i] = round(sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value,2)
#                     # print(finArmDiam[i])
#             elif cell.value == "Fin. Arm radius [m]":
#                 for i in range(0, numberLink):
#                     finArmRadius[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value 
#                     # print(finArmRadius[i])
#             elif cell.value == "Mass [g]":
#                 for i in range(0, numberLink):
#                     mass[i] = round(sheetLT.cell(row=cell.row+1 + i, column=cell.column ).value,2)
#                     # print(mass[i])
#             # elif cell.value == "Angle":
#             #     for i in range(0, numberLink):
#             #         angle[i] = sheetLT.cell(row=cell.row+1+i, column=cell.column ).value
#             #         # print(angle[i])

#     winding =  {
#         'numberLink':numberLink,
        
#     }
#     return winding

# def upload_file_winding(request):
#     if request.method == 'POST' and 'file' in request.FILES:
#         winding = handle_uploaded_file_winding(request.FILES['file'])
#         print(winding)
#         if 'error' in winding:
#             return JsonResponse(winding, status=400)
#         return JsonResponse(winding)
#     return JsonResponse({'error': 'Invalid request'}, status=400)
