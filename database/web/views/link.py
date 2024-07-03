import warnings
from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django import forms
from django.forms import modelformset_factory
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Frame
from django import forms

from web import models
import django_filters

class LinkFilter(django_filters.FilterSet):
    interface1 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)
    interface2 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)

    class Meta:
        model = models.Link
        fields = {
            'project': ['exact'], 
            'part': ['exact'],
            'linkName': ['icontains'],
        }

def link_list(request):
    """ list of link """
    links = None
    link_filter = LinkFilter(request.GET, queryset=models.Link.objects.all())

    message = "No interface to display. Please use the filter to load data."

    if any(request.GET.values()):
        links = link_filter.qs
        message = "No data found."
    elif 'filter' in request.GET:
        links = link_filter.qs
        message = "No data found."
    
    return render(request, 'link/link_list.html', {'filter': link_filter, 'links': links, 'message':message})



def link_valid(request):
    """ list of link """
    links = None
    link_filter = LinkFilter(request.GET, queryset=models.Link.objects.filter(part__valid=True))

    message = "No interface to display. Please use the filter to load data."

    if any(request.GET.values()):
        links = link_filter.qs
        message = "No data found."
    elif 'filter' in request.GET:
        links = link_filter.qs
        message = "No data found."
    
    return render(request, 'link/link_valid.html', {'filter': link_filter, 'links': links, 'message':message})


def link_input(request):
    """ list of link """

    # [obj,]
    queryset = models.Link.objects.all().order_by("id")
  
    return render(request, 'link/link_input.html', {"queryset": queryset})

   
class LinkModelForm(forms.ModelForm):
    interface1 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)
    interface2 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)


    class Meta:
        model = models.Link
        fields = [
                  
                  'linkName',  'interface1', 'interface2', 
                'sequence', 'ratio', 'armDiam', 'armSection', 'cycle', 'angle',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field_object in self.fields.items():
            field_object.widget.attrs = {"class": "form-control"}


class ProjectPartForm(forms.Form):
    project = forms.ModelChoiceField(queryset=models.Project.objects.all(), required=True)
    part = forms.ModelChoiceField(queryset=models.Part.objects.all(), required=True)
    
# class InterfaceForm(forms.Form):
#     interface1 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)
#     interface2 = forms.ModelChoiceField(queryset=models.Interface.objects.all(), required=True)

LinkFormSet = modelformset_factory(
    models.Link,
    form=LinkModelForm,
    extra=1 
)

def link_add_multiple(request):
    projects = models.Project.objects.all()
    parts = models.Part.objects.all()
    # interfaces = models.Interface.objects.all()
    formset = LinkFormSet(queryset=models.Link.objects.none())
    project_part_form = ProjectPartForm()
    # interface_form = InterfaceForm()

    if request.method == 'POST':
        print("post")
        formset = LinkFormSet(request.POST)
        project_part_form = ProjectPartForm(request.POST)
        # interface_form = InterfaceForm(request.POST)

        if project_part_form.is_valid(): 
            print("project")
            project = project_part_form.cleaned_data['project']
            part = project_part_form.cleaned_data['part']
            # if interface_form.is_valid():
            #     print("interface")
            #     interface1 = interface_form.cleaned_data['interface1']
            #     interface2 = interface_form.cleaned_data['interface2']
            if formset.is_valid() : 
                print("formset")
                instances = formset.save(commit=False)
                
                for form,instance in zip(formset.forms,instances):
                    interface1 = form.cleaned_data['interface1']
                    interface2 = form.cleaned_data['interface2']
                    instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                    instance.project = project
                    instance.part = part
                    instance.interface1 = interface1
                    instance.interface2 = interface2
                    instance.save()
                return redirect('/link/list/')  # Replace with your redirect URL
            else:
                # Handle formset or project_part_form validation errors here
                print(formset.errors)
            # else:
            #     # Handle formset or project_part_form validation errors here
            #     print(interface_form.errors)
        else:
            # Handle formset or project_part_form validation errors here
            print(project_part_form.errors)

    return render(request, 'link/link_add_multiple.html', {
        'projects': projects,
        'parts': parts,
        # 'interfaces': interfaces,
        'formset': formset,
        'project_part_form': project_part_form,
        # 'interface_form': interface_form,
        'linkError': "Link name is Required",
    })

def link_modify_multiple(request):
    # Get filtered data
    link_filter = LinkFilter(request.GET, queryset=models.Link.objects.all())
    
    # Define form set
    LinkFormSet = modelformset_factory(models.Link, form=LinkModelForm, extra=0)
    
    formset = None
    message = "No link to display. Please use the filter to load data."


    if request.method == 'POST':
        formset = LinkFormSet(request.POST)
        if formset.is_valid():
            # formset.save()
            instances = formset.save(commit=False)
            
            for instance in instances:
                instance.user = models.User.objects.filter(id=request.info_dict['id']).first()  
                instance.save()
            return redirect('/link/list/')
        
    else:
        if any(request.GET.values()):
            formset = LinkFormSet(queryset=link_filter.qs)
            if not link_filter.qs.exists():
                message = "No data found."
        elif 'filter' in request.GET:
            formset = LinkFormSet(queryset=link_filter.qs)
            if not link_filter.qs.exists():
                message = "No data found."

    return render(request, 'link/link_modify_multiple.html', {
        'filter': link_filter,
        'formset': formset,
        # 'interfaces': interfaces,
        'linkError': "Link name is Required",
        'message': message,
    })

def link_add(request):
    if request.method == "GET":
        form = LinkModelForm()
        return render(request, 'link/link_form.html', {"form": form})

    form = LinkModelForm(data=request.POST)
    if not form.is_valid():
        return render(request, 'link/link_form.html', {"form": form})


    # form.save()
    link = form.save(commit=False)
    link.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    link.save()
    return redirect('/link/list/')


class LinkEditModelForm(forms.ModelForm):
    class Meta:
        model = models.Link
        fields = ['linkName', 
                # 'length', 
                'sequence', 'ratio', 
                'armDiam', 'armSection', 
                'cycle',  
                # 'finArmSection', 'finArmDiam', 'finArmRadius', 'mass', 
                'angle',]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        for name, filed_object in self.fields.items():
            filed_object.widget.attrs = {"class": "form-control"}


def link_edit(request, aid):
    link_object = models.Link.objects.filter(id=aid).first()

    if request.method == "GET":
        form = LinkEditModelForm(instance=link_object)
        return render(request, 'link/link_form.html', {"form": form})

    form = LinkEditModelForm(instance=link_object, data=request.POST)
    if not form.is_valid():
        return render(request, 'link/link_form.html', {"form": form})

    # form.save()
    link = form.save(commit=False)
    link.user = models.User.objects.filter(id=request.info_dict['id']).first()  
    link.save()

    return redirect('/link/list/')


def link_delete(request):
    aid = request.GET.get("aid")
    # models.Link.objects.filter(id=aid).delete()
    link = models.Link.objects.filter(id=aid).first()
    if link:
        link.user = models.User.objects.filter(id=request.info_dict['id']).first()  
        link.delete()
    
    return JsonResponse({"status": True})

def link_delete_mult(request):
    aid = request.GET.get("aid")
    if not aid:
        return JsonResponse({"status": False, "error": "ID cannot be empty"})

    try:
        link = models.Fiber.objects.get(id=aid)
        # link.delete()
        if link:
            link.user = models.User.objects.filter(id=request.info_dict['id']).first()  
            link.delete()

        return JsonResponse({"status": True})
    except models.Fiber.DoesNotExist:
        return JsonResponse({"status": False, "error": "Fiber not found"})

def custom_warning_filter(message, category, filename, lineno, file=None, line=None):
    if "Unknown extension is not supported and will be removed" in str(message) or \
       "Conditional Formatting extension is not supported and will be removed" in str(message):
        return
    else:
        warnings.showwarning(message, category, filename, lineno, file, line)

# Apply custom warning filter
warnings.showwarning = custom_warning_filter


def handle_uploaded_file_link(f):
    # load Excel document
    wb = load_workbook(filename=f, data_only=True)
    if 'CoverPage' not in wb.sheetnames:
        return {'error': 'CoverPage sheet not found'}
    if 'Link Table' not in wb.sheetnames:
        return {'error': 'Link Table sheet not found'}
    # if 'Sequence' not in wb.sheetnames:
    #     return {'error': 'Sequence sheet not found'}
    
    sheetCP = wb['CoverPage']
    sheetLT = wb['Link Table']
    sheetSE = wb['Sequence']

    numberLink = 0

    link_data = {}
    linkName = {}
    interface1 = {}
    interface2 = {}

    armDiam = {}
    armSection = {}

    cycle = {}
    sequence = {}
    ratio= {}
    mass = {}
    angle = {}
    
    project_data = {}
    for row in sheetCP.iter_rows():
        for cell in row:
            if cell.value == "Program : ":
                project_data['program'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Customer : ":
                project_data['customer'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
            elif cell.value == "Project No:":
                project_data['projectNo'] = sheetCP.cell(row=cell.row, column=cell.column + 2).value
    project_data['projectName'] = project_data['projectNo'] +' - '+ project_data['customer'] +' - '+ project_data['program']

    try:
        project = models.Project.objects.get(projectName=project_data['projectName'])
        link_data['project_id'] = project.id
    except models.Project.DoesNotExist:
            return {'error': 'Project not found'}
    try:
        part = models.Part.objects.get(partName=project_data['projectName'])
        link_data['part_id'] = part.id
    except models.Part.DoesNotExist:
            return {'error': 'Part not found'}
    
    
               
    for row in sheetLT.iter_rows():
        for cell in row:
            if cell.value == "Link":
                i = 1
                while sheetLT.cell(row=cell.row+i, column=cell.column ).value != None:
                    numberLink += 1
                    # print(numberLink)
                    i +=1
                # print(numberLink)

        for cell in row:
            if cell.value == "Link":
                for i in range(0, numberLink):
                    # print(cell.row, cell.column)
                    linkName[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
                    # print(str(i) + linkName[i])
            elif cell.value == "Bushing Bi":
                for i in range(0, numberLink):
                    interface_obj = models.Interface.objects.filter(interfaceName=sheetLT.cell(row=cell.row+ 1 + i, column=cell.column).value, part_id=part.id).first()
                    if interface_obj:
                        interface_id = interface_obj.id
                        interface1[i] = interface_id
                    else:
                        interface1[i] = None
                    
                    
                    # print(interface1[i])
            elif cell.value == "Bushing Bj":
                for i in range(0, numberLink):
                    interface_obj = models.Interface.objects.filter(interfaceName=sheetLT.cell(row=cell.row+ 1 + i, column=cell.column).value, part_id=part.id).first()
                    if interface_obj:
                        interface_id = interface_obj.id
                        interface2[i] = interface_id
                    else:
                        interface2[i] = None
                    # print(interface2[i])
            elif cell.value == "Sequence":
                for i in range(0, numberLink):
                    sequence[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column).value
                    # print(sequence[i])
           
            elif cell.value == "Arm diam. [mm]":
                for i in range(0, numberLink):
                    armDiam[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
                    # print(armDiam[i])
            elif cell.value == "Arm Sec. [mm²]":
                for i in range(0, numberLink):
                    armSection[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
                    # print(armSection[i])
            elif cell.value == "Cycle #":
                for i in range(0, numberLink):
                    cycle[i] = sheetLT.cell(row=cell.row+ 1 + i, column=cell.column ).value
                    # print(cycle[i])
           
            elif cell.value == "Angle":
                for i in range(0, numberLink):
                    angle[i] = sheetLT.cell(row=cell.row+1+i, column=cell.column ).value
                    # print(angle[i])
            else :
                for i in range(0, numberLink):
                    angle[i] = 0

    if sheetSE:
        for row in sheetSE.iter_rows():
            for cell in row:
                if cell.value == "Ratio":
                    for i in range(0, numberLink):
                        # print(cell.row, cell.column)
                        ratio[i] = sheetSE.cell(row=cell.row+ 1 + i, column=cell.column ).value
                        # print(str(i) + linkName[i])
    else:
        for i in range(0, numberLink):
            ratio[i] = "--"

    links = []
                                                 
    # for i in range(numberLink):
    #     try:
    #         interface1_instance = models.Interface.objects.filter(interfaceName=interface1[i], part_id=part.id).first().id
    #         interface2_instance = models.Interface.objects.filter(interfaceName=interface2[i], part_id=part.id).first().id
    #     except models.Interface.DoesNotExist:
    #         return {'error': f'Interface not found for link {linkName[i]}'}
        

    link =  {
        'numberLink':numberLink,
        'link_data':link_data,
        'linkName':linkName,
        # 'interface1':interface1_instance ,
        # 'interface2':interface2_instance ,
        'interface1': interface1,
        'interface2': interface2,
        # 'lengthLink' :lengthLink,
        'armDiam':armDiam,
        'armSection':armSection,
        'cycle':cycle,
        'sequence':sequence,
        'ratio':ratio,
        # 'finArmSection':finArmSection,
        # 'finArmDiam':finArmDiam,
        # 'finArmRadius':finArmRadius,
        # 'mass':mass,
        'angle':angle,
    }
    return link

def upload_file_link(request):
    if request.method == 'POST' and 'file' in request.FILES:
        link = handle_uploaded_file_link(request.FILES['file'])
        # print(link)
        if 'error' in link:
            return JsonResponse(link, status=400)
        return JsonResponse(link)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def draw_rotated_header(canvas, x, y, text, width, height, angle):
    canvas.saveState()
    canvas.translate(x, y)
    canvas.rotate(angle)
    p = Paragraph(text, getSampleStyleSheet()['Normal'])
    p.wrapOn(canvas, height, width)  # width and height are swapped because of rotation
    p.drawOn(canvas, 0, 0)
    canvas.restoreState()

def format_value(value):
        return str(value) if value else "--"

def download_links_pdf(request):
    f = LinkFilter(request.GET, queryset=models.Link.objects.filter(part__valid=True))
    links = f.qs

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="links.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    styles = getSampleStyleSheet()

    p.setFont("Helvetica-Bold", 12)
    p.drawString(30, height - 40, "Links List")      
    p.setFont("Helvetica-Bold", 10)
    p.drawString(150, height - 40, "\"--\" means the value is None")
    p.setDash() 
    p.setLineWidth(1)
    p.line(30, height - 50, width - 30, height - 50)
    p.setFont("Helvetica-Bold", 10)
    y_start = height - 70
    x_offset = 30
    y_offset = 30  
    
            
    headers = ["Link Name", "Interface1","Interface2","Length [mm]","Sequence","Arm diam. [mm]",
                "Arm Sec. [mm²]","Cycle #","Fin. Arm Sec. [mm²]","Fin. Arm diam. [mm]",
                "Fin. Arm radius [m]","Mass [g]","Angle" ]
    column_widths = [35, 30, 30, 40, 150, 35, 
                     35, 20, 35, 35, 
                     50, 35, 30]

    current_project = None
    current_part = None

    p.setFont("Helvetica", 7)
    y = y_start

    for index, link in enumerate(links):
        next_is_title = (index + 1 < len(links)) and (
            links[index + 1].project.projectName != link.project.projectName or 
            links[index + 1].part.partName != link.part.partName
        )

        if link.project.projectName != current_project or link.part.partName != current_part:
            if current_project is not None:
                y -= 10  
                p.setDash() 
                p.setLineWidth(1)
                p.line(x_offset, y, width - x_offset, y)

            current_project = link.project.projectName
            current_part = link.part.partName
            y -= y_offset

            p.setFont("Helvetica-Bold", 10)
            p.drawString(x_offset, y, "Project Name: " + current_project)
            y -= y_offset/2
            p.drawString(x_offset, y, "Part      Name: " + current_part)
            y -= y_offset/2
            p.setDash() 
            p.setLineWidth(1)
            p.line(x_offset, y, width - x_offset, y)
            y -= y_offset/2

            for i, header in enumerate(headers):
                draw_rotated_header(p, x_offset + sum(column_widths[:i]) + column_widths[i] / 2, y - 3 * y_offset, header, 200,200,  90)
            y -= 4 * y_offset


            p.setFont("Helvetica", 9)

        
        p.setDash(1, 2) 
        p.setLineWidth(1)
        values = [
            format_value(link.linkName), format_value(link.interface1), format_value(link.interface2), format_value(link.length),
            Paragraph(format_value(link.sequence), styles['Normal']), format_value(link.armDiam), format_value(link.armSection), format_value(link.cycle),
            format_value(link.finArmSection), format_value(link.finArmDiam), format_value(link.finArmRadius), format_value(link.mass),
            format_value(link.angle)
        ]

        p.drawString(x_offset, y, values[0])
        for i in range(1, len(values)):
            p.line(x_offset + sum(column_widths[:i]) - 5, y + y_offset / 2, x_offset + sum(column_widths[:i]) - 5, y - y_offset / 2)
            if isinstance(values[i], Paragraph):
                values[i].wrapOn(p, column_widths[i]-10, y_offset)
                values[i].drawOn(p, x_offset + sum(column_widths[:i]), y - y_offset / 2)
            else:
                p.drawString(x_offset + sum(column_widths[:i]), y, values[i])

        y -= y_offset
        if next_is_title:
            p.showPage()
            p.setFont("Helvetica-Bold", 12)
            p.drawString(30, height - 40, "Links List")
            p.setFont("Helvetica-Bold", 10)
            p.drawString(150, height - 40, "\"--\" means the value is None")
            p.setDash() 
            p.setLineWidth(1)
            p.line(30, height - 50, width - 30, height - 50)
            p.setFont("Helvetica-Bold", 10)
            y = height - 70
            current_project = None
            current_part = None 
        else:
            if y < 50 :
                p.showPage()
                p.setFont("Helvetica-Bold", 12)
                p.drawString(30, height - 40, "Links List (continued)")
                p.setFont("Helvetica-Bold", 10)
                p.drawString(200, height - 40, "\"--\" means the value is None")
                p.setDash() 
                p.setLineWidth(1)
                p.line(30, height - 50, width - 30, height - 50)
                p.setFont("Helvetica-Bold", 10)
                y = height - 70
                current_project = None
                current_part = None 

    p.showPage()
    p.save()
    return response
